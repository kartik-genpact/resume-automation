import nltk
import string
import pandas as pd
from nltk import word_tokenize
from nltk.corpus import stopwords
import os
import matplotlib.pyplot as plt
# import sys, fitz
from zipfile import ZipFile
import docx2txt
import streamlit as st
import plotly.graph_objects as go
# import shutil
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import re
import pymongo
import random
from docx2python import docx2python
# import dns


class ResumeSummarizer():
    def __init__(self):
        nltk.download('stopwords')
        nltk.download('punkt')
        global writer
        writer = pd.ExcelWriter("excel_report.xlsx", engine='xlsxwriter')
        pass

    # def pdf_to_text(self, file): # Converting PDF to text
    #     doc = fitz.open(file)
    #     text = ''
    #     for page in doc:
    #         text = text + str(page.get_text())
    #     text = ' '.join(text.split('\n'))
    #     return text

    def docx_to_text(self, file): #convert docx to text
        text = docx2txt.process(file)
        text = ' '.join(text.split('\n'))
        text = ' '.join(text.split('\t'))
        return text

    def preprocess(self, text): #preprocess the text
      #re.sub('[^a-zA-Z]',' ',text)
      text = text.lower()
      text_p = "".join([char for char in text if char not in string.punctuation])
      words = word_tokenize(text_p)
      stop_words = stopwords.words('english')
      filtered_words = [word for word in words if word not in stop_words]
      result = ' '.join(filtered_words)
      return result

    def db_create(self, result, file_path): #create a dataframe of skills for a employee
        template = pd.read_excel('template1.xlsx')
        skill_list = template.columns.tolist()
        skill_dict = {}
        skill_count = []
        skill_count_dict = {}
        for i in range(len(skill_list)):
            skill_dict[skill_list[i]] = list(template[skill_list[i]].dropna())
        col = ['Candidate_Name']+skill_list
        final_db = pd.DataFrame(columns=col)
        row=[]
        base = os.path.basename(file_path)
        filename = os.path.splitext(base)[0]
        if '_' in filename:
            name = filename.split('_')
        elif '-' in filename:
            name = filename.split('-')
        else:
            name = filename.split(' ')
        name2 = name[0]
        row.append(name2)
        def countOccurrences(str, word):
            wordslist = list(str.split())
            return wordslist.count(word)
        for i in range(len(skill_list)):
            skill_count.append(skill_list[i] + '_count')
        for i in range(len(skill_count)):
            skill_count_dict[skill_count[i]] = 0
        for i in range(len(skill_list)):
            for j in skill_dict[skill_list[i]]:
                k = countOccurrences(result, j)
                skill_count_dict[skill_count[i]] = k + skill_count_dict[skill_count[i]]
            row.append(skill_count_dict[skill_count[i]])
        final_db.loc[len(final_db.index)] = row
        final_db = final_db.infer_objects()
        return final_db

    def isfloat(self, value): # checks wheter a string is float or not
        try:
            float(value)
            return True
        except ValueError:
            return False

    # def plot(self, final_db):
    #     ax = final_db.plot.barh(title="Resume categorization by skills", legend=True, stacked=True)
    #     labels = []
    #     for j in final_db.columns:
    #         for i in final_db.index:
    #             label = str(j)+": " + str(final_db.loc[i][j])
    #             labels.append(label)
    #     patches = ax.patches
    #     for label, rect in zip(labels, patches):
    #         width = rect.get_width()
    #         if width > 0:
    #             x = rect.get_x()
    #             y = rect.get_y()
    #             height = rect.get_height()
    #             ax.text(x + width/2., y + height/2., label, ha='center', va='center')
    #     ax.figure.savefig('summary2.png')
    #     plt.show()
    #     #st.pyplot(ax)

    def unzip(self, files): #unzips the zip folder
        with ZipFile(files, 'r') as zipObj:
           zipObj.extractall('output')

    def parse_single_file(self): #returns single files fron the zip folder
        files = []
        pwd = os.getcwd()
        wd = pwd + '/output'
        swd = wd+'/'
        directories = os.listdir(wd)
        for file in directories:
            files.append(swd+file)
        return files

    def create_excel_table(self, df, sheet_name): #creates an excel table formatting
        worksheet = writer.sheets[sheet_name]
        (max_row, max_col) = df.shape
        column_settings = []
        for header in df.columns:
            column_settings.append({'header': header})
        worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})
        worksheet.set_column(0, max_col - 1, 12)

    def check(self, word, lst): #checks if a word is in list or not
        if word in lst:
            return True
        else:
            return False

    # def get_experience(self, zip_file): #get the experience along with the UI and graphs
    #     st.success('...Done!')
    #     st.header("Experience ")
    #     st.write('Total work experience of each employee')
    #     self.unzip(zip_file)
    #     files = self.parse_single_file()
    #     col=['Candidate_Name', 'Experience(yrs)']
    #     experience_db = pd.DataFrame(columns=col)
    #     for file in files:
    #         if (file[-3:] == 'pdf'):
    #             text = self.pdf_to_text(file)
    #         else:
    #             text = self.docx_to_text(file)
    #         text = re.sub('[^.a-zA-Z0-9]', ' ', text)
    #         row = []
    #         base = os.path.basename(file)
    #         filename = os.path.splitext(base)[0]
    #         if '_' in filename:
    #             name = filename.split('_')
    #         elif '-' in filename:
    #             name = filename.split('-')
    #         else:
    #             name = filename.split(' ')
    #         name2 = name[0]
    #         row.append(name2)
    #         k = (text.split())
    #         count = 0
    #         l = ['experience', 'experience.', 'Exp.', 'exp.', 'exp', 'Exp']
    #         for i in range(len(l)):
    #             if self.check(l[i], k):
    #                 t = k.index(l[i])
    #                 if t > 5:
    #                     for i in range(t, t - 5, -1):
    #                         if k[i].isnumeric() or self.isfloat(k[i]):
    #                             if self.isfloat(k[i]):
    #                                 row.append(float(k[i]))
    #                                 count = 1
    #                             else:
    #                                 row.append(int(k[i]))
    #                                 count = 1
    #                 elif t<5:
    #                     for i in range(t, t + 5):
    #                         if k[i].isnumeric() or self.isfloat(k[i]):
    #                             if self.isfloat(k[i]):
    #                                 row.append(float(k[i]))
    #                                 count = 1
    #                             else:
    #                                 row.append(int(k[i]))
    #                                 count = 1
    #         if count == 0:
    #             row.append(0)
    #         # print(row)
    #         experience_db.loc[len(experience_db.index)] = row
    #     global ex_db_v
    #     ex_db_v = experience_db
    #     ex_db_v = ex_db_v.set_index('Candidate_Name')
    #     experience_db.to_excel(writer, sheet_name='Experience', startrow=1, header=False, index=False)
    #     self.create_excel_table(experience_db, 'Experience')
    #     # writer.save()
    #     st.dataframe(ex_db_v)
    #     st.bar_chart(ex_db_v)
    #     return experience_db

    #def get_location(self):

    def flatten_df(self, df):
        db = df
        for i in range(len(df.columns)):
            for j in range(len(df[i])):
                db[i][j] = df[i][j][0][0]
        return db

    def get_info(self, file):
        document = docx2python(file)
        d = dict()
        l = self.flatten_df(pd.DataFrame(document.body_runs[1]))
        for i in range(len(l[0])):
            l[1][i] = l[1][i].lower()
        l[1][2] = float(l[1][2])
        d['personal_info_df'] = l
        d['summary_list'] = ((document.body_runs[2])[0][0][2:])
        t = self.flatten_df(pd.DataFrame(document.body_runs[3]))
        for i in range(len(t[0])):
            t[0][i] = t[0][i].lower()
            t[1][i] = float(t[1][i])
        d['skills_df'] = t
        d['domain'] = (document.body_runs[4][0][0][2])
        # d['project_df'] = flatten_df(pd.DataFrame(document.body_runs[5]))
        d['education_df'] = self.flatten_df(pd.DataFrame(document.body_runs[9]))
        d['certification_list'] = (document.body_runs[10][0][0][2:])
        d['visa_df'] = self.flatten_df(pd.DataFrame(document.body_runs[11]).drop(2, axis=1))
        d['link'] = str(document.body_runs[12][0][0][1][0])  #[9:-4]
        return d

    def get_experience_db(self, zip_file): #get the experience data table only
        st.header("Experience")
        st.write("Total experience of all employees")
        col = ['Candidate_Name', 'Experience(yrs)']
        experience_db = pd.DataFrame(columns=col)
        name, exp = [], []
        if zip_file==None:
            pass
        elif zip_file.name[-3:]=='zip':
            st.info('Extracting and Processing the resumes...')
            st.success('...Done')
            self.unzip(zip_file)
            files = self.parse_single_file()
            for file in files:
                u = self.get_info(file)
                p = u['personal_info_df']
                name.append(p.loc[p[0] == 'Name', 1].item())
                exp.append(p.loc[p[0] == 'Years of Experience', 1].item())
                experience_db['Candidate_Name'] = (name)
                experience_db['Experience(yrs)'] = (exp)
                experience_db.to_excel(writer, sheet_name='Experience', startrow=1, header=False, index=False)
                self.create_excel_table(experience_db, 'Experience')
                s = experience_db
                s = s.set_index('Candidate_Name')
                st.dataframe(s)
                st.bar_chart(s)
                return experience_db
        elif zip_file.name[-4:]=='docx':
            st.info('Extracting and Processing the resumes...')
            st.success('...Done')
            u = self.get_info(zip_file)
            p = u['personal_info_df']
            name.append(p.loc[p[0] == 'Name', 1].item())
            exp.append(p.loc[p[0] == 'Years of Experience', 1].item())
            experience_db['Candidate_Name'] = (name)
            experience_db['Experience(yrs)'] = (exp)
            experience_db.to_excel(writer, sheet_name='Experience', startrow=1, header=False, index=False)
            self.create_excel_table(experience_db, 'Experience')
            s = experience_db
            s = s.set_index('Candidate_Name')
            st.dataframe(s)
            st.bar_chart(s)
            return experience_db

    def get_just_experience_db(self, zip_file):
        self.unzip(zip_file)
        files = self.parse_single_file()
        col = ['Candidate_Name', 'Experience(yrs)']
        experience_db = pd.DataFrame(columns=col)
        name, exp = [], []
        for file in files:
            u = self.get_info(file)
            p = u['personal_info_df']
            name.append(p.loc[p[0] == 'Name', 1].item())
            exp.append(p.loc[p[0] == 'Years of Experience', 1].item())
        experience_db['Candidate_Name'] = (name)
        experience_db['Experience(yrs)'] = (exp)
        return experience_db

    def summarize(self, zip_file): #summarization of everything
        if zip_file==None:
            pass
        elif (zip_file.name)[-3:]=='zip':
            self.unzip(zip_file)
            files = self.parse_single_file()
            dbs=[]
            for file in files:
                text = self.docx_to_text(file)
                result = self.preprocess(text)
                final_db = self.db_create(result, file)
                dbs.append(final_db)
            excel_db = dbs
            dfs = [df.set_index('Candidate_Name') for df in dbs]
            db = (pd.concat(dfs))
            excel_db = pd.concat(excel_db)
            excel_db.to_excel(writer, sheet_name='Skills', startrow=1, header=False, index=False)
            self.create_excel_table(excel_db, 'Skills')
            skills = pd.DataFrame(columns=['skill', 'count'])
            skills['skill'] = db.columns.tolist()
            count = []
            for i in list(db):
                k = db[i].tolist()
                count.append(len([1 for i in k if i > 0]))
            skills['count'] = count
            # st.success('...Done!')
            st.header("Tabular Representation of Skills ")
            st.write("The score generated for each employee for the respective skill:")
            st.dataframe(db)
            st.header("Graphical Representation of Skills")
            st.write("Easy to handpick employee according to a particular skill set: ")
            st.bar_chart(db)
            fig = go.Figure(
                go.Pie(
                    labels=skills['skill'],
                    values=skills['count'],
                    hoverinfo="label+percent",
                    textinfo="value+label",
                    hole=0.2
                ))
            # self.plot(db)
            # st.image('summary.png')
            st.header("Distribution of Employee Count by Skill")
            st.write("No of employees having the particular skill :")
            sv2 = skills
            skills.to_excel(writer, sheet_name='Skill vs employee', startrow=1, header=False, index=False)
            self.create_excel_table(skills, 'Skill vs employee')
            writer.save()
            wb = load_workbook('test/excel_report.xlsx')
            charts = wb.create_sheet('charts', 0)
            active = wb['charts']
            new_skill = skills
            new_skill = new_skill[~(new_skill == 0).any(axis=1)]
            plot = new_skill.groupby(['skill']).sum().plot.pie(y='count', autopct='%1.0f%%', legend=False,
                                                               title='Employee Count vs Skill')
            plot.figure.savefig('summary.png')
            plt.show()
            active.add_image(Image('summary.png'), 'A1')
            wb.save('excel_report.xlsx')
            sv2 = sv2.set_index('skill')
            st.dataframe(sv2)
            st.plotly_chart(fig)
            # shutil.rmtree(r'output')
            return excel_db
        elif zip_file.name[-4:]=='docx':
            dbs = []
            text = self.docx_to_text(zip_file)
            result = self.preprocess(text)
            final_db = self.db_create(result, zip_file.name)
            dbs.append(final_db)
            excel_db = dbs
            dfs = [df.set_index('Candidate_Name') for df in dbs]
            db = (pd.concat(dfs))
            excel_db = pd.concat(excel_db)
            excel_db.to_excel(writer, sheet_name='Skills', startrow=1, header=False, index=False)
            self.create_excel_table(excel_db, 'Skills')
            skills = pd.DataFrame(columns=['skill', 'count'])
            skills['skill'] = db.columns.tolist()
            count = []
            for i in list(db):
                k = db[i].tolist()
                count.append(len([1 for i in k if i > 0]))
            skills['count'] = count
            # st.success('...Done!')
            st.header("Tabular Representation of Skills ")
            st.write("The score generated for each employee for the respective skill:")
            st.dataframe(db)
            st.header("Graphical Representation of Skills")
            st.write("Easy to handpick employee according to a particular skill set: ")
            st.bar_chart(db)
            fig = go.Figure(
                go.Pie(
                    labels=skills['skill'],
                    values=skills['count'],
                    hoverinfo="label+percent",
                    textinfo="value+label",
                    hole=0.2
                ))
            # self.plot(db)
            # st.image('summary.png')
            st.header("Distribution of Employee Count by Skill")
            st.write("No of employees having the particular skill :")
            sv2 = skills
            skills.to_excel(writer, sheet_name='Skill vs employee', startrow=1, header=False, index=False)
            self.create_excel_table(skills, 'Skill vs employee')
            writer.save()
            wb = load_workbook('test/excel_report.xlsx')
            charts = wb.create_sheet('charts', 0)
            active = wb['charts']
            new_skill = skills
            new_skill = new_skill[~(new_skill == 0).any(axis=1)]
            plot = new_skill.groupby(['skill']).sum().plot.pie(y='count', autopct='%1.0f%%', legend=False,
                                                               title='Employee Count vs Skill')
            plot.figure.savefig('summary.png')
            plt.show()
            active.add_image(Image('summary.png'), 'A1')
            wb.save('excel_report.xlsx')
            sv2 = sv2.set_index('skill')
            st.dataframe(sv2)
            st.plotly_chart(fig)
            # shutil.rmtree(r'output')
            return excel_db

    def skill_db(self, zip_file): #return the skill db
        if zip_file.name[-3:]=='zip':
            self.unzip(zip_file)
            files = self.parse_single_file()
            dbs = []
            for file in files:
                text = self.docx_to_text(file)
                result = self.preprocess(text)
                final_db = self.db_create(result, file)
                dbs.append(final_db)
        elif zip_file[-4:]=='docx':
            dbs = []
            text = self.docx_to_text(zip_file)
            result = self.preprocess(text)
            final_db = self.db_create(result, zip_file)
            dbs.append(final_db)
        excel_db = dbs
        dfs = [df.set_index('Candidate_Name') for df in dbs]
        db = (pd.concat(dfs))
        excel_db = pd.concat(excel_db)
        return excel_db

    # def filter_current(self, zip_file): #filter the current uploaded data
    #     experience = self.get_experience_db(zip_file)
    #     skill = self.skill_db(zip_file)
    #     skill['Experience'] = list(experience['Experience(yrs)'])
    #     skill_2 = skill.drop('Candidate_Name', axis = 1)
    #     skill_list = list(skill_2.columns)
    #     st.header('Filter Employees')
    #     st.write('Filter Employees based on Skill and Years of experience')
    #     primary_skill = str(st.selectbox('Primary Skill:', skill_list))
    #     secondary_skill = str(st.selectbox('Secondary Skill:', skill_list))
    #     yrs_of_exp = int(st.slider('Years of Experience:', 1, 10, 2))
    #     # filter_skill_db = skill.filter('@primary_skill>5')
    #     filter_skill_db = skill[(skill[primary_skill]>2) & (skill[secondary_skill]>1)]
    #     filter_skill_db = filter_skill_db[filter_skill_db['Experience']==yrs_of_exp]
    #     if st.button("Filter"):
    #         for i in list(filter_skill_db['Candidate_Name']):
    #             st.subheader(i)

    # def mongodb_upload_dataframe(self, df): #upload dataframe to mongodb
    #     try:
    #         client = pymongo.MongoClient(
    #             "mongodb+srv://admin:pavilion15@cluster0.pefyq.mongodb.net/myFirstDatabase?retryWrites=true&w=majority")
    #         db = client['ResumeData']
    #         col = db['resumes']
    #         dic = df.to_dict(orient='list')
    #         col.insert_many([dic])
    #         print('Uploaded to MongoDB successfully!')
    #     except:
    #         print("Upload failed!")

    def mongodb_view(self): #view mongodb db
        client = pymongo.MongoClient(
            "mongodb+srv://admin:pavilion15@cluster0.pefyq.mongodb.net/myFirstDatabase?retryWrites=true&w=majority")
        db = client['ResumeData']
        col = db['resumes']
        masterdb = pd.DataFrame(columns=['name', 'exp', 'status'])
        view = col.find()
        choice = st.radio('What do you want to do?', ['View Database', 'Change Status of Employee'])
        if choice=='View Database':
            masterdb = pd.DataFrame(columns=['Name', 'OHR ID', 'Experience', 'Position', 'Location',
                                             'Site', 'Contact', 'Skills', 'Status', 'Resume Link'])
            for i in view:
                row = []
                row.append(i['Name'])
                row.append((i['OHR ID']))
                row.append(i['Years of Experience'])
                row.append(i['Position'])
                row.append(i['Location'])
                row.append(i['Site(Onsite/Remote)'])
                row.append(i['Contact'])
                row.append(i['Skill'])
                row.append(i['status'])
                row.append(i['link'])
                masterdb.loc[len(masterdb.index)] = row
            st.write(masterdb.to_html(escape=False, index=False), unsafe_allow_html=True)
        elif choice=='Change Status of Employee':
            namelist = []
            for i in view:
                namelist.append(i['Name'])
            n = st.selectbox('Select the Candidate', namelist)
            status = st.selectbox('Select the status', ['Selected-Extended', 'Proposed', 'Selected-New', 'Blocked', 'Unassigned', 'Unavailable - Resigned',
                      'Unavailable - Cost Transfer', 'Unavailable - Maternity Leave', 'Unavailable - Sabbatical', 'Unavailable - Extended Maternity Leave'])
            if st.button('Update Status'):
                col.update_one({'Name': n }, { "$set": { 'status': status } })

    def mongodb_upload(self, zip_file):
        #extract and upload data to db
        client = pymongo.MongoClient(
            "mongodb+srv://admin:pavilion15@cluster0.pefyq.mongodb.net/myFirstDatabase?retryWrites=true&w=majority")
        db = client['ResumeData']
        col = db['resumes']
        if zip_file==None:
            pass
        elif zip_file.name[-3:]=='zip':
            self.unzip(zip_file)
            files = self.parse_single_file()
            for file in files:
                u = (self.get_info(file))
                p = pd.concat([u['personal_info_df'], u['skills_df'], u['education_df'], u['visa_df']])
                dt = pd.Series(['status', 'Unassigned'], index=p.columns)
                p = p.append(dt, ignore_index=True)
                l = pd.Series(['link', u['link']], index=p.columns)
                p = p.append(l, ignore_index=True)
                skill = []
                for i in range(len(u['skills_df'][0])):
                    o = []
                    o.append(u['skills_df'][0][i])
                    o.append(u['skills_df'][1][i])
                    skill.append(o)
                sl = pd.Series(['Skill', skill], index=p.columns)
                p = p.append(sl, ignore_index=True)
                dic_list = dict(p.values)
                view = col.find_one({'OHR ID': p[1][1]})
                if view == None:
                    col.insert_one(dic_list)
            st.header('Excel Report:')
            st.write(
                "Download a Report in Excel format, which comprises of all the detailed analysis and summary of each resume.")
            with open("excel_report.xlsx", "rb") as file:
                btn = st.download_button(
                    label="Download Excel Report",
                    data=file,
                    file_name="excel_report.xlsx")
        elif zip_file.name[-4:]=='docx':
            u = (self.get_info(zip_file))
            p = pd.concat([u['personal_info_df'], u['skills_df'], u['education_df'], u['visa_df']])
            dt = pd.Series(['status', 'Unassigned'], index=p.columns)
            p = p.append(dt, ignore_index=True)
            l = pd.Series(['link', u['link']], index=p.columns)
            p = p.append(l, ignore_index=True)
            skill = []
            for i in range(len(u['skills_df'][0])):
                o = []
                o.append(u['skills_df'][0][i])
                o.append(u['skills_df'][1][i])
                skill.append(o)
            sl = pd.Series(['Skill', skill], index=p.columns)
            p = p.append(sl, ignore_index=True)
            dic_list = dict(p.values)
            view = col.find_one({'OHR ID': p[1][1]})
            if view == None:
                col.insert_one(dic_list)
            st.header('Excel Report:')
            st.write(
                "Download a Report in Excel format, which comprises of all the detailed analysis and summary of each resume.")
            with open("excel_report.xlsx", "rb") as file:
                btn = st.download_button(
                    label="Download Excel Report",
                    data=file,
                    file_name="excel_report.xlsx")

    def mongodb_allskills(self): #gets all the skills from database
        client = pymongo.MongoClient(
            "mongodb+srv://admin:pavilion15@cluster0.pefyq.mongodb.net/myFirstDatabase?retryWrites=true&w=majority")
        db = client['ResumeData']
        col = db['resumes']
        find = col.find()
        master = []
        for i in find:
            all = (i['Skill'])
            for j in all:
                master.append(j[0])
        return list(set(master))

    def mongodb_summarize(self): #summarization from the database
        choice = st.selectbox('What do you want to do ?', ['Filter', 'Skill Count'])
        skill_list = self.mongodb_allskills()
        client = pymongo.MongoClient(
            "mongodb+srv://admin:pavilion15@cluster0.pefyq.mongodb.net/myFirstDatabase?retryWrites=true&w=majority")
        db = client['ResumeData']
        col = db['resumes']

        if choice=='Filter':
            st.header('Filter Employees')
            st.write('Filter Employees based on Skill and Years of experience')
            col1, col2 = st.columns([2, 1])
            primary_skill = str(col1.selectbox('Primary Skill:', skill_list))
            ps_exp = (col2.number_input('Skill Experience', step = 0.5, value = 2.5))
            if ps_exp-int(ps_exp)==0:
                ps_exp = int(ps_exp)
            secondary_skill = str(col1.selectbox('Secondary Skill:', skill_list))
            ss_exp = (col2.number_input('Skill Experience ', step = 0.5, value=2.5))
            if ss_exp-int(ss_exp)==0:
                ss_exp = int(ss_exp)
            location = str(col1.text_input('Base Location')).lower()
            site = str(col2.radio('Site: ', ['Onsite', 'Remote'])).lower()
            visa_status = str(col1.radio('Visa Status', ['Available', 'Unavailable']))
            yrs_of_exp = (col2.number_input('Total Years of Experience:', step = 0.5, value = 2.5))
            if yrs_of_exp-int(yrs_of_exp)==0:
                yrs_of_exp = int(yrs_of_exp)
            status = str(st.radio('Status:',['Selected-Extended', 'Proposed', 'Selected-New', 'Blocked', 'Unassigned', 'Unavailable - Resigned',
                      'Unavailable - Cost Transfer', 'Unavailable - Maternity Leave', 'Unavailable - Sabbatical', 'Unavailable - Extended Maternity Leave']))
            view = col.find({'$and':[{primary_skill: {"$lte": ps_exp}}, {secondary_skill: {'$lte': ss_exp}}, {'Location': location},
                                     {'Site(Onsite/Remote)': site}, {'Visa 1': visa_status}, {'status': status}, {'Years of Experience': {'$lte': yrs_of_exp}}]})
            if st.button('Filter'):
                masterdb = pd.DataFrame(columns=['Name', 'OHR ID', 'Experience', 'Position', 'Location',
                                                 'Site', 'Contact', 'Skills', 'Status', 'Resume Link'])
                for i in view:
                    row = []
                    row.append(i['Name'])
                    row.append((i['OHR ID']))
                    row.append(i['Years of Experience'])
                    row.append(i['Position'])
                    row.append(i['Location'])
                    row.append(i['Site(Onsite/Remote)'])
                    row.append(i['Contact'])
                    row.append(i['Skill'])
                    row.append(i['status'])
                    row.append(i['link'])
                    masterdb.loc[len(masterdb.index)] = row
                st.write(masterdb.to_html(escape=False, index=False), unsafe_allow_html=True)
        elif choice=='Skill Count':
            master = []
            find = col.find()
            for i in find:
                all = (i['Skill'])
                for j in all:
                    master.append(j[0])
            skill = list(set(master))
            count = []
            for i in range(len(skill)):
                count.append(master.count(skill[i]))
            sd = pd.DataFrame(columns=['skill', 'count'])
            sd['skill'] = skill
            sd['count'] = count
            st.dataframe(sd)
            new_skill = sd[~(sd == 0).any(axis=1)]
            fig = go.Figure(
                go.Pie(
                    labels=new_skill['skill'],
                    values=new_skill['count'],
                    hoverinfo="label+percent",
                    textinfo="value+label",
                    hole=0.2
                ))
            st.plotly_chart(fig)

obj = ResumeSummarizer()
# obj.get_experience_db('Resumesnew.zip')
# obj.mongodb_upload('Resumesnew.zip')

# obj = ResumeSummarizer()
# st.set_page_config(initial_sidebar_state='expanded',layout='wide')
# st.sidebar.image('mainlogo.png')
# st.image('mainlogo.png')
# st.image(['tagline1.png', 'tagline2.png'] )
# st.title("Internal Employee Detail Summarizer")
# choice = st.sidebar.radio("What's on your mind ?", ['New Summarization', 'Summarize from DB', 'View DB'])
# if choice=='New Summarization':
#     global zip_file
#     zip_file = st.sidebar.file_uploader('Upload the Resume ZIP file: ')
#     def generate_summary():
#         obj.get_experience_db(zip_file)
#         (obj.summarize(zip_file))
#         obj.mongodb_upload(zip_file)
#     if st.sidebar.button('Start Analysis'):
#         st.info('Extracting and Processing the resumes...')
#         generate_summary()
#         st.header('Excel Report:')
#         st.write("Download a Report in Excel format, which comprises of all the detailed analysis and summary of each resume.")
#         with open("excel_report.xlsx", "rb") as file:
#              btn = st.download_button(
#                  label="Download Excel Report",
#                  data=file,
#                  file_name="excel_report.xlsx")
# elif choice=='View DB':
#     obj.mongodb_view()
# elif choice=='Summarize from DB':
#     obj.mongodb_summarize()
